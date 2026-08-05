# -*- coding: utf-8 -*-
import os
import re
import json
import traceback
import csv
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import subprocess
import requests
import uvicorn
from dotenv import load_dotenv
from passlib.hash import bcrypt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, BackgroundTasks, Request, Depends, HTTPException, status, Form
from fastapi.responses import HTMLResponse, RedirectResponse

# ================= 加载环境变量 =================
# 自动定位项目根目录下的 .env 文件
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
env_path = os.path.join(BASE_DIR, ".env")

if os.path.exists(env_path):
    load_dotenv(dotenv_path=env_path, override=True)
else:
    load_dotenv(dotenv_path=r"d:\sshAuto\.env", override=True)

FEISHU_APP_ID = os.getenv("FEISHU_APP_ID")
FEISHU_APP_SECRET = os.getenv("FEISHU_APP_SECRET")
TARGET_RECEIVER = os.getenv("TARGET_RECEIVER")

ADMIN_USER = os.getenv("ADMIN_USER", "admin")
ADMIN_PASS = os.getenv("ADMIN_PASS")
ADMIN_PASS_HASH = os.getenv("ADMIN_PASS_HASH")
COOKIE_NAME = os.getenv("COOKIE_NAME", "current_admin_session")
SESSION_TOKEN_SECRET = os.getenv("SESSION_TOKEN_SECRET", "default_secret_key_2026")

app = FastAPI()

# ================= 统一路径定义 & 自动创建 =================
CSV_DIR = os.path.join(BASE_DIR, "reports_csv")
XLSX_DIR = os.path.join(BASE_DIR, "reports_xlsx")
IMG_DIR = os.path.join(BASE_DIR, "reports_imgs")
TEMP_DIR = os.path.join(BASE_DIR, "tempFlies")

for d in [CSV_DIR, XLSX_DIR, IMG_DIR, TEMP_DIR]:
    os.makedirs(d, exist_ok=True)

plt.rcParams['font.sans-serif'] = ['SimHei']  
plt.rcParams['axes.unicode_minus'] = False     

def _visual_len(s):
    length = 0
    for ch in str(s):
        code = ord(ch)
        if (0x4E00 <= code <= 0x9FFF or 0x3000 <= code <= 0x303F or
                0xFF00 <= code <= 0xFFEF or 0x2E80 <= code <= 0x2EFF):
            length += 2
        else:
            length += 1
    return length

# ================= NVR 专属渲染函数 =================
def convert_nvr_to_png(csv_path, png_path):
    if not os.path.exists(csv_path): return False
    try:
        df = pd.read_csv(csv_path, encoding="utf-8-sig")
        if "所属NVR" in df.columns:
            m = df["所属NVR"] == df["所属NVR"].shift()
            df.loc[m, "所属NVR"] = ""
    except Exception as e:
        print(f"[后台控制台] 读取 NVR CSV 失败: {e}")
        return False

    headers = list(df.columns)
    row_count = len(df)
    header_line_count = 1
    LINE_UNIT_HEIGHT = 0.38
    BASE_ROW_PAD = 0.25
    fig_width = 17  
    fig_height = 1.2 + header_line_count * 0.5 + (row_count * LINE_UNIT_HEIGHT + BASE_ROW_PAD)
    
    fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=200)
    ax.axis('off')

    col_widths_raw = {}
    for col in headers:
        max_w = _visual_len(col)
        for val in df[col].fillna("-"):
            max_w = max(max_w, _visual_len(val))
        col_widths_raw[col] = max_w + 3

    total_w = sum(col_widths_raw.values())
    col_widths = [col_widths_raw[col] / total_w for col in headers]

    tb = ax.table(
        cellText=df.fillna("-").values, colLabels=df.columns,
        cellLoc='center', loc='center',
        colColours=['#1F4E79'] * len(headers), colWidths=col_widths  
    )

    tb.auto_set_font_size(False)
    tb.set_fontsize(10)

    total_line_units = row_count + 1
    for row_idx in range(row_count + 1):
        for col_idx in range(len(headers)):
            tb[row_idx, col_idx].set_height(1 / total_line_units)

    for i in range(len(headers)):
        cell = tb[0, i]
        cell.get_text().set_color('white')
        cell.get_text().set_weight('bold')

    for row_idx in range(1, row_count + 1):
        status_val = str(df.iloc[row_idx-1, headers.index("状态")]).strip()
        if status_val == "正常在线": bg_color, text_color = '#C6EFCE', '#006100'  
        elif status_val == "部分不可达": bg_color, text_color = '#FFF2CC', '#B25E00'  
        else: bg_color, text_color = '#FFC6C6', '#9C0006'  

        for col_idx in range(len(headers)):
            cell = tb[row_idx, col_idx]
            cell.set_facecolor(bg_color)
            cell.get_text().set_color(text_color)

    plt.title("📊 科伦特监控状态一览表", fontsize=15, weight='bold', pad=12, color='#1F4E79')
    plt.savefig(png_path, bbox_inches='tight', dpi=200)
    plt.close()
    return True

def convert_nvr_to_xlsx(csv_path, xlsx_path):
    if not os.path.exists(csv_path): return False
    try: df = pd.read_csv(csv_path, encoding="utf-8-sig")
    except Exception as e: return False

    df = df.fillna("-")
    headers = list(df.columns)
    wb = Workbook()
    ws = wb.active
    ws.title = "监控状态一览"

    FONT_NAME = "微软雅黑"
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")  
    GREEN_FILL, YELLOW_FILL, RED_FILL = PatternFill("solid", fgColor="C6EFCE"), PatternFill("solid", fgColor="FFF2CC"), PatternFill("solid", fgColor="FFC6C6")
    GREEN_TEXT, YELLOW_TEXT, RED_TEXT = "006100", "B25E00", "9C0006"
    THIN = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="📊 科伦特监控状态一览表")
    title_cell.font = Font(name=FONT_NAME, size=14, bold=True, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    for c, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=col_name)
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.border = BORDER

    for r_offset, (_, row) in enumerate(df.iterrows()):
        r = 3 + r_offset
        status_val = str(row.get("状态", "")).strip()
        bg, text_color = (GREEN_FILL, GREEN_TEXT) if status_val == "正常在线" else ((YELLOW_FILL, YELLOW_TEXT) if status_val == "部分不可达" else (RED_FILL, RED_TEXT))

        for c, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c, value=str(row[col_name]).strip())
            cell.font = Font(name=FONT_NAME, size=10, color=text_color)
            cell.fill = bg
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for c, col_name in enumerate(headers, start=1):
        max_w = max([_visual_len(val) for val in df[col_name]] + [_visual_len(col_name)])
        ws.column_dimensions[get_column_letter(c)].width = max_w + 5

    wb.save(xlsx_path)
    return True

# ================= 交换机核心巡检通用渲染 =================
def convert_txt_to_png(csv_path, png_path):
    if not os.path.exists(csv_path): return False
    try:
        df = pd.read_csv(csv_path, encoding="utf-8-sig")
    except:
        return False
        
    headers = list(df.columns)
    ALARM_COL = "最新关键告警"
    WRAP_WIDTH = 26

    def wrap_by_visual_width(text, width=WRAP_WIDTH):
        if not text or str(text) in ["无", "-", "无告警", "nan"]: return "正常 (无告警)"
        text = str(text)
        lines, cur, cur_w = [], "", 0
        for ch in text:
            ch_w = 2 if _visual_len(ch) == 2 else 1
            if cur_w + ch_w > width and cur:
                lines.append(cur)
                cur, cur_w = "", 0
            cur += ch
            cur_w += ch_w
        if cur: lines.append(cur)
        return "\n".join(lines)

    if ALARM_COL in df.columns:
        df[ALARM_COL] = df[ALARM_COL].apply(wrap_by_visual_width)

    row_line_counts = []
    for _, row in df.iterrows():
        lines_in_row = str(row[ALARM_COL]).count("\n") + 1 if ALARM_COL in df.columns else 1
        row_line_counts.append(lines_in_row)

    header_line_count = 1
    total_line_units = header_line_count + sum(row_line_counts)
    LINE_UNIT_HEIGHT = 0.34
    BASE_ROW_PAD = 0.30
    fig_width = 16
    fig_height = 1.1 + header_line_count * 0.5 + sum((lc * LINE_UNIT_HEIGHT + BASE_ROW_PAD) for lc in row_line_counts)
    
    fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=200)
    ax.axis('off')

    col_widths_raw = {}
    for col in headers:
        max_w = _visual_len(col)
        for val in df[col]:
            max_w = max(max_w, max([_visual_len(line) for line in str(val).split("\n")]) if col == ALARM_COL else _visual_len(val))
        col_widths_raw[col] = max_w + 2

    total_w = sum(col_widths_raw.values())
    col_widths = [col_widths_raw[col] / total_w for col in headers]

    tb = ax.table(cellText=df.values, colLabels=df.columns, cellLoc='center', loc='center', colColours=['#3498db'] * len(headers), colWidths=col_widths)
    tb.auto_set_font_size(False)
    tb.set_fontsize(9.5)

    for row_idx in range(len(df) + 1):
        lines = header_line_count if row_idx == 0 else row_line_counts[row_idx - 1]
        for col_idx in range(len(headers)):
            tb[row_idx, col_idx].set_height(lines / total_line_units)

    for i in range(len(headers)):
        cell = tb[0, i]
        cell.get_text().set_color('white')
        cell.get_text().set_weight('bold')

    for row_idx in range(1, len(df) + 1):
        for col_idx in range(len(headers)):
            cell = tb[row_idx, col_idx]
            val = str(df.iloc[row_idx-1, col_idx]).strip()
            col_name = headers[col_idx]
            bg_color, text_color = '#ffffff', '#000000'
            cell.set_text_props(multialignment='center')

            if "CPU" in col_name and "%" in val:
                try:
                    if float(val.replace('%', '')) > 80: bg_color = '#FFC6C6'; text_color = '#9C0006'
                except: pass
            elif "内存" in col_name and "%" in val:
                try:
                    mem_num = float(re.findall(r'\d+\.?\d*', val)[0])
                    if mem_num > 85: bg_color = '#FFC6C6'; text_color = '#9C0006'
                except: pass
            elif "状态" in col_name:
                if "正常" in val: bg_color = '#C6EFCE'; text_color = '#006100'
                elif "提示" in val: bg_color = '#FFF2CC'; text_color = '#B25E00'
                elif "异常" in val: bg_color = '#FFC6C6'; text_color = '#9C0006'
            elif "告警" in col_name or "日志" in col_name:
                status_val = str(df.iloc[row_idx-1, headers.index("状态")]).strip()
                if "正常" in status_val: bg_color = '#C6EFCE'; text_color = '#006100'
                elif "提示" in status_val: bg_color = '#FFF2CC'; text_color = '#B25E00'
                elif "异常" in status_val: bg_color = '#FFD2D2'; text_color = '#9C0006'

            cell.set_facecolor(bg_color)
            cell.get_text().set_color(text_color)

    plt.title("📊 终端实时巡检及日志告警看板", fontsize=14, weight='bold', pad=8)
    plt.savefig(png_path, bbox_inches='tight', dpi=200)
    plt.close()
    return True

def convert_csv_to_xlsx(csv_path, xlsx_path):
    if not os.path.exists(csv_path): return False
    try: df = pd.read_csv(csv_path, encoding="utf-8-sig")
    except: return False
    
    headers = list(df.columns)
    ALARM_COL = "最新关键告警"
    STATUS_COL = "状态"
    ALARM_WRAP_WIDTH = 40

    wb = Workbook()
    ws = wb.active
    ws.title = "巡检看板"

    FONT_NAME = "Arial"
    HEADER_FILL = PatternFill("solid", fgColor="3498DB")
    GREEN_FILL, YELLOW_FILL, RED_FILL, RED_FILL_ALARM = PatternFill("solid", fgColor="C6EFCE"), PatternFill("solid", fgColor="FFF2CC"), PatternFill("solid", fgColor="FFC6C6"), PatternFill("solid", fgColor="FFD2D2")
    GREEN_TEXT, YELLOW_TEXT, RED_TEXT = "006100", "B25E00", "9C0006"
    THIN = Side(style="thin", color="B7B7B7")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="📊 终端实时巡检及日志告警看板")
    title_cell.font = Font(name=FONT_NAME, size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    for c, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=col_name)
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.border = BORDER

    for r_offset, (_, row) in enumerate(df.iterrows()):
        r = 3 + r_offset
        status_val = str(row.get(STATUS_COL, "")).strip()
        alarm_raw = str(row.get(ALARM_COL, "")).strip() if ALARM_COL in headers else ""
        alarm_display = "正常 (无告警)" if alarm_raw in ["", "无", "-", "无告警", "nan"] else alarm_raw

        for c, col_name in enumerate(headers, start=1):
            val = alarm_display if col_name == ALARM_COL else str(row[col_name]).strip()
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left" if col_name in ["设备IP", "设备名称", ALARM_COL] else "center", vertical="center", wrap_text=(col_name == ALARM_COL))

            bg, text_color = None, None
            if "CPU" in col_name and "%" in val:
                try:
                    if float(val.replace("%", "")) > 80: bg, text_color = RED_FILL, RED_TEXT
                except: pass
            elif "内存" in col_name and "%" in val:
                m = re.findall(r"\d+\.?\d*", val)
                if m and float(m[0]) > 85: bg, text_color = RED_FILL, RED_TEXT
            elif col_name == STATUS_COL:
                if "正常" in val: bg, text_color = GREEN_FILL, GREEN_TEXT
                elif "提示" in val: bg, text_color = YELLOW_FILL, YELLOW_TEXT
                elif "异常" in val: bg, text_color = RED_FILL, RED_TEXT
            elif "告警" in col_name or "日志" in col_name:
                if "正常" in status_val: bg, text_color = GREEN_FILL, GREEN_TEXT
                elif "提示" in status_val: bg, text_color = YELLOW_FILL, YELLOW_TEXT
                elif "异常" in status_val: bg, text_color = RED_FILL_ALARM, RED_TEXT

            if bg:
                cell.fill = bg
                cell.font = Font(name=FONT_NAME, size=10, color=text_color)

        needed_lines = max(1, -(-_visual_len(alarm_display) // ALARM_WRAP_WIDTH))
        ws.row_dimensions[r].height = 18 if needed_lines <= 1 else 15 * needed_lines

    for c, col_name in enumerate(headers, start=1):
        max_w = max([_visual_len(val) for val in df[col_name]] + [_visual_len(col_name)])
        ws.column_dimensions[get_column_letter(c)].width = ALARM_WRAP_WIDTH + 4 if col_name == ALARM_COL else max_w + 4

    ws.freeze_panes = f"A4"
    wb.save(xlsx_path)
    return True

# ================= 飞书 API 连接 & 发信组件 =================
def get_feishu_tenant_access_token():
    url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
    try:
        res = requests.post(url, json={"app_id": FEISHU_APP_ID, "app_secret": FEISHU_APP_SECRET}, timeout=10)
        return res.json().get("tenant_access_token")
    except Exception as e:
        return None

def send_feishu_message(text_content, receive_id=TARGET_RECEIVER):
    token = get_feishu_tenant_access_token()
    if not token: return
    url = "https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type=chat_id"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json; charset=utf-8"}
    escaped_text = text_content.replace('\\', '\\\\').replace('"', '\\"').replace('\n', '\\n')
    payload = {"receive_id": receive_id, "msg_type": "text", "content": f'{{"text": "{escaped_text}"}}'}
    try: requests.post(url, headers=headers, json=payload, timeout=10)
    except: pass

def upload_and_send_feishu_image(image_path, receive_id=TARGET_RECEIVER):
    token = get_feishu_tenant_access_token()
    if not token: return
    upload_url = "https://open.feishu.cn/open-apis/im/v1/images"
    headers = {"Authorization": f"Bearer {token}"}
    try:
        with open(image_path, "rb") as f:
            files = {"image_type": (None, "message"), "image": (os.path.basename(image_path), f, "image/png")}
            res = requests.post(upload_url, headers=headers, files=files, timeout=15)
            image_key = res.json().get("data", {}).get("image_key")
    except: return

    if not image_key: return
    send_url = "https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type=chat_id"
    send_headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json; charset=utf-8"}
    payload = {"receive_id": receive_id, "msg_type": "image", "content": f'{{"image_key": "{image_key}"}}'}
    requests.post(send_url, headers=send_headers, json=payload, timeout=10)

def upload_and_send_feishu_file(file_path, mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", receive_id=TARGET_RECEIVER):
    token = get_feishu_tenant_access_token()
    if not token: return
    upload_url = "https://open.feishu.cn/open-apis/im/v1/files"
    headers = {"Authorization": f"Bearer {token}"}
    try:
        with open(file_path, "rb") as f:
            files = {"file_type": (None, "stream"), "file_name": (None, os.path.basename(file_path)), "file": (os.path.basename(file_path), f, mime_type)}
            res = requests.post(upload_url, headers=headers, files=files, timeout=20)
            file_key = res.json().get("data", {}).get("file_key")
    except: return

    if not file_key: return
    send_url = "https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type=chat_id"
    send_headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json; charset=utf-8"}
    payload = {"receive_id": receive_id, "msg_type": "file", "content": f'{{"file_key": "{file_key}"}}'}
    requests.post(send_url, headers=send_headers, json=payload, timeout=10)

# ================= 核心任务执行分发引擎 =================
def run_network_script(script_name: str, receive_id: str = TARGET_RECEIVER):
    script_path = os.path.join(BASE_DIR, script_name)
    try:
        if script_name != "test.py":
            send_feishu_message(f"⏳ 任务 {script_name} 已开始运行，请稍候...", receive_id=receive_id)
        
        temp_file_path = os.path.join(TEMP_DIR, "temp_table.txt")
        temp_csv_path = os.path.join(CSV_DIR, "temp_report.csv")
        nvr_temp_csv = os.path.join(CSV_DIR, "nvr_temp_report.csv")
        
        report_image_path = os.path.join(IMG_DIR, "report.png")
        report_xlsx_path = os.path.join(XLSX_DIR, "report.xlsx")
        nvr_image_path = os.path.join(IMG_DIR, "科伦特监控状态一览表.png")
        nvr_xlsx_path = os.path.join(XLSX_DIR, "科伦特监控状态一览表.xlsx")
        
        python_executable = os.path.join(BASE_DIR, ".venv", "Scripts", "python.exe")
        if not os.path.exists(python_executable):
            python_executable = "python"

        result = subprocess.run(
            [python_executable, script_path, receive_id],
            capture_output=True, text=True, encoding="gbk", errors="ignore", timeout=600, cwd=BASE_DIR
        )

        if result.returncode == 0:
            if "test" in script_name:
                print("👉 [主服务控制台] 异步执行 test.py 成功。")
                return
            elif "sshAuto" in script_name:
                if os.path.exists(temp_csv_path):
                    if convert_txt_to_png(temp_csv_path, report_image_path):
                        upload_and_send_feishu_image(report_image_path, receive_id=receive_id)
                    if convert_csv_to_xlsx(temp_csv_path, report_xlsx_path):
                        upload_and_send_feishu_file(report_xlsx_path, receive_id=receive_id)
                    return
                send_feishu_message(f"⚠️ {script_name} 运行成功，但未捕获到可视化报表数据。", receive_id=receive_id)
            elif "NVR_devices" in script_name:
                if os.path.exists(nvr_temp_csv):
                    if convert_nvr_to_png(nvr_temp_csv, nvr_image_path):
                        upload_and_send_feishu_image(nvr_image_path, receive_id=receive_id)
                    if convert_nvr_to_xlsx(nvr_temp_csv, nvr_xlsx_path):
                        upload_and_send_feishu_file(nvr_xlsx_path, receive_id=receive_id)
                else:
                    send_feishu_message(f"⚠️ {script_name} 运行成功，但没有产生 nvr_temp_report.csv 报表文件。", receive_id=receive_id)
            else:
                send_feishu_message(f"✅ {script_name} 运行成功！拓扑图已生成在备份目录。", receive_id=receive_id)
        else:
            err_msg = f"❌ {script_name} 运行失败。返回值: {result.returncode}\n报错:\n{result.stderr[-300:]}"
            send_feishu_message(err_msg, receive_id=receive_id)
    except Exception as e:
        send_feishu_message(f"❌ 系统运行异常: {str(e)}", receive_id=receive_id)

# ================= 身份验证拦截器 =================
def get_current_user(request: Request):
    """检查 Cookie 中是否存在合法的 Session 凭证，若没有则直接重定向到登录页"""
    session_cookie = request.cookies.get(COOKIE_NAME)
    if session_cookie != SESSION_TOKEN_SECRET:
        raise HTTPException(
            status_code=status.HTTP_303_SEE_OTHER, 
            headers={"Location": "/login"}
        )
    return session_cookie

# ================= 登录页面路由 =================
@app.get("/login", response_class=HTMLResponse)
def login_page():
    return """
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>网管系统 - 登录</title>
        <script src="https://cdn.tailwindcss.com"></script>
    </head>
    <body class="bg-slate-900 min-h-screen flex items-center justify-center p-4">
        <div class="bg-white/10 backdrop-blur-md p-8 rounded-2xl shadow-2xl w-full max-w-md border border-white/20">
            <h2 class="text-2xl font-bold text-white text-center mb-6">🔒 移动控制终端登录</h2>
            <form action="/login" method="post" class="space-y-4">
                <div>
                    <label class="block text-sm font-medium text-slate-300 mb-1">账号</label>
                    <input type="text" name="username" required class="w-full px-4 py-2.5 rounded-xl bg-slate-800 border border-slate-700 text-white focus:outline-none focus:border-blue-500">
                </div>
                <div>
                    <label class="block text-sm font-medium text-slate-300 mb-1">密码</label>
                    <input type="password" name="password" required class="w-full px-4 py-2.5 rounded-xl bg-slate-800 border border-slate-700 text-white focus:outline-none focus:border-blue-500">
                </div>
                <button type="submit" class="w-full py-3 bg-blue-600 hover:bg-blue-700 text-white font-semibold rounded-xl transition duration-200 shadow-lg active:scale-[0.98] mt-2">
                    立即登录
                </button>
            </form>
        </div>
    </body>
    </html>
    """

# ================= 登录验证接口（处理表单提交） =================
@app.post("/login")
def do_login(username: str = Form(...), password: str = Form(...)):
    """验证账号与密码（兼顾明文与哈希），验证成功颁发 Session 凭证"""
    is_user_valid = (username == ADMIN_USER)
    is_pass_valid = False

    # 1. 优先验证明文密码 (ADMIN_PASS)
    if ADMIN_PASS and password == ADMIN_PASS:
        is_pass_valid = True
    # 2. 其次校验哈希密码 (ADMIN_PASS_HASH)
    elif ADMIN_PASS_HASH:
        try:
            clean_hash = ADMIN_PASS_HASH.strip("'\"")  # 过滤可能误带的单双引号
            if bcrypt.verify(password, clean_hash):
                is_pass_valid = True
        except Exception as e:
            print(f"👉 [Hash校验异常] {e}")

    if is_user_valid and is_pass_valid:
        response = RedirectResponse(url="/", status_code=status.HTTP_303_SEE_OTHER)
        response.set_cookie(
            key=COOKIE_NAME, 
            value=SESSION_TOKEN_SECRET, 
            httponly=True,  # 禁止前端 JavaScript 读取 Cookie，防 XSS 攻击
            secure=True,    # 仅允许在 HTTPS 连接中传输 Cookie
            samesite="lax", # 防止跨站请求伪造 (CSRF)
            path="/"
        )
        return response
    
    return HTMLResponse(
        content="<script>alert('账号或密码错误，请重新输入！'); window.location.href='/login';</script>", 
        status_code=400
    )

# ================= 权限受控：现代控制台前端页面 =================
@app.get("/", response_class=HTMLResponse)
def mobile_control_panel(current_user: str = Depends(get_current_user)):
    return """
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>网管移动控制终端</title>
        <script src="https://cdn.tailwindcss.com"></script>
        <style>
            .glass-card {
                background: rgba(255, 255, 255, 0.95);
                backdrop-filter: blur(10px);
                border: 1px solid rgba(255, 255, 255, 0.2);
            }
        </style>
    </head>
    <body class="bg-slate-100 min-h-screen flex items-center justify-center p-4 relative">
        
        <!-- 右上角退出登录按钮 -->
        <div class="absolute top-4 right-4">
            <a href="/logout" class="flex items-center space-x-1 px-3 py-1.5 bg-white hover:bg-red-50 text-slate-600 hover:text-red-600 font-medium text-xs rounded-lg shadow-sm border border-slate-200 transition duration-200">
                <span>🚪 退出登录</span>
            </a>
        </div>

        <div class="glass-card w-full max-w-md p-8 rounded-2xl shadow-xl mt-8">
            <h1 class="text-2xl font-bold text-slate-800 mb-2 text-center">📱 网管移动控制终端</h1>
            <p class="text-slate-500 text-sm mb-6 text-center">自动化运维脚本集中调度看板</p>
            <hr class="mb-6 border-slate-200">
            
            <div class="space-y-4">
                <button onclick="runTask('test')" class="block w-full py-4 px-4 bg-orange-500 hover:bg-orange-600 text-white font-semibold rounded-xl transition duration-200 text-center shadow-md active:scale-[0.98]">
                    ⚡ 快速测试服务 (test.py)
                </button>
                <button onclick="runTask('inspect')" class="block w-full py-4 px-4 bg-emerald-500 hover:bg-emerald-600 text-white font-semibold rounded-xl transition duration-200 text-center shadow-md active:scale-[0.98]">
                    🚀 开始网络巡检 (sshAuto.py)
                </button>
                <button onclick="runTask('nvr')" class="block w-full py-4 px-4 bg-purple-500 hover:bg-purple-600 text-white font-semibold rounded-xl transition duration-200 text-center shadow-md active:scale-[0.98]">
                    📹 监控设备检查 (NVR_devices.py)
                </button>
                <button onclick="runTask('topology')" class="block w-full py-4 px-4 bg-blue-500 hover:bg-blue-600 text-white font-semibold rounded-xl transition duration-200 text-center shadow-md active:scale-[0.98]">
                    📊 生成网络拓扑 (draw_topology.py)
                </button>
            </div>
        </div>

        <script>
            function runTask(type) {
                fetch('/run-task?task_type=' + type)
                .then(res => {
                    if (res.status === 401 || res.status === 418) {
                        window.location.href = '/login';
                        return;
                    }
                    return res.json();
                })
                .then(data => {
                    if(data) alert(data.message);
                });
            }
        </script>
    </body>
    </html>
    """

# ================= 权限受控：退出登录接口 =================
@app.get("/logout")
def do_logout():
    """清除登录 Cookie 并重定向回登录页面"""
    response = RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    response.delete_cookie(key=COOKIE_NAME, path="/")
    return response

# ================= 权限受控：后台指令派发 =================
@app.get("/run-task")
def trigger_task(task_type: str, background_tasks: BackgroundTasks, current_user: str = Depends(get_current_user)):
    if task_type == "test": background_tasks.add_task(run_network_script, "test.py")
    elif task_type == "topology": background_tasks.add_task(run_network_script, "draw_topology.py")
    elif task_type == "inspect": background_tasks.add_task(run_network_script, "sshAuto.py")
    elif task_type == "nvr": background_tasks.add_task(run_network_script, "NVR_devices.py")
    return {"message": "正在后台派发处理指令..."}

# ================= 公开区域：飞书事件总线 Webhook（不受拦截影响） =================
@app.post("/webhook")
async def feishu_webhook(request: Request, background_tasks: BackgroundTasks):
    # 💡 拦截飞书网络超时后的自动重试推送，防止重复拉起 NVR/巡检脚本
    if request.headers.get("X-Lark-Retry-Reason") or request.headers.get("X-Lark-Retry-Nonce"):
        print("⚠️ 收到飞书重试消息，已自动忽略，防止重复触发脚本。")
        return {"code": 0, "msg": "duplicate retry ignored"}

    body = await request.json()
    print(f"收到飞书请求: {body}") 
    if body.get("type") == "url_verification": 
        return {"challenge": body.get("challenge")}
    
    header = body.get("header", {})
    if header.get("event_type") == "im.message.receive_v1":
        event = body.get("event", {})
        message = event.get("message", {})
        chat_id = message.get("chat_id")
        
        if message.get("message_type") == "text":
            try:
                content_json = json.loads(message.get("content", "{}"))
                clean_text = re.sub(r'@[^ ]+\s*', '', content_json.get("text", "")).strip().lower()
            except: clean_text = ""

            if any(k in clean_text for k in ["测试", "ping", "test", "在线"]): background_tasks.add_task(run_network_script, "test.py", chat_id)
            elif "巡检" in clean_text or "inspect" in clean_text:
                background_tasks.add_task(run_network_script, "sshAuto.py", chat_id)
                send_feishu_message("🤖 收到！已为您开启核心网络设备巡检任务，请耐心等待看板送达...", receive_id=chat_id)
            elif "拓扑" in clean_text or "topology" in clean_text: background_tasks.add_task(run_network_script, "draw_topology.py", chat_id)
            elif "监控" in clean_text or "nvr" in clean_text:
                background_tasks.add_task(run_network_script, "NVR_devices.py", chat_id)
                send_feishu_message("🤖 收到！已为您开启 NVR 监控状态排查，请稍候...", receive_id=chat_id)
    return {"code": 0, "msg": "success"}

if __name__ == "__main__":
    print(f"👉 [系统启动] 读取到的管理员账号: {ADMIN_USER}")
    print(f"👉 [系统启动] 明文密码: {'已配置' if ADMIN_PASS else '未配置'}, Hash密码: {'已配置' if ADMIN_PASS_HASH else '未配置'}")
    
    cert_key = os.path.join(BASE_DIR, "26178675_paynehe.me_nginx", "paynehe.me.key")
    cert_pem = os.path.join(BASE_DIR, "26178675_paynehe.me_nginx", "paynehe.me.pem")

    uvicorn.run("agent_server:app", host="0.0.0.0", port=8000, 
                ssl_keyfile=cert_key if os.path.exists(cert_key) else r"D:\sshAuto\26178675_paynehe.me_nginx\paynehe.me.key", 
                ssl_certfile=cert_pem if os.path.exists(cert_pem) else r"D:\sshAuto\26178675_paynehe.me_nginx\paynehe.me.pem")